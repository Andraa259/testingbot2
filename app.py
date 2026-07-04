import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import matplotlib.pyplot as plt
import io
import re

# 1. Konfigurasi Dasar Halaman Web
st.set_page_config(page_title="Psikometri Auto-Analyzer & Plotter", layout="wide")

st.title("📊 Automated Psychometric Plotter & Analyzer (v4.6 - Evaluasi 1)")
st.write("Sistem otomatisasi Kolom O dan pembuatan Grafik Kuadrant Berdasarkan Aturan Evaluasi 1 Dosen.")

# 2. Komponen Input (Gateway)
uploaded_file = st.file_uploader("Upload File Excel 'hasil analisis quiz kls A.xlsx'", type=["xlsx"])

if uploaded_file is not None:
    try:
        # Load workbook asli untuk mempertahankan seluruh data, rumus, dan format
        file_bytes = uploaded_file.read()
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
        
        if 'RANGKUMAN' not in wb.sheetnames:
            st.error("Error: Sheet bernama 'RANGKUMAN' tidak ditemukan!")
        else:
            ws_rangkuman = wb['RANGKUMAN']
            
            # Membaca data sekunder lewat pandas untuk kebutuhan kalkulasi & koordinat grafik
            df_calc = pd.read_excel(io.BytesIO(file_bytes), sheet_name='RANGKUMAN', header=0)
            
            # Helper untuk merapikan nama opsi (A, B, & C)
            def format_opsi_nama(daftar_opsi):
                if not daftar_opsi:
                    return ""
                if len(daftar_opsi) == 1:
                    return daftar_opsi[0]
                if len(daftar_opsi) == 2:
                    return f"{daftar_opsi[0]} & {daftar_opsi[1]}"
                return ", ".join(daftar_opsi[:-1]) + f", & {daftar_opsi[-1]}"

            # --- PROSES 1: OTOMASI ANALISIS DISTRAKTOR (KOLOM O) ---
            with st.spinner("Sistem sedang memproses Kolom O berbasis warna kuning KJ..."):
                for idx in range(1, len(df_calc)):
                    row = df_calc.iloc[idx]
                    excel_row_num = idx + 2
                    
                    try:
                        mean_val = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0.0
                        if mean_val >= 1.0:
                            ws_rangkuman.cell(row=excel_row_num, column=15, value="Semua Distraktor tidak efektif")
                            continue
                            
                        opsi_pct = {
                            'A': float(row.iloc[7]) if pd.notna(row.iloc[7]) else 0.0,
                            'B': float(row.iloc[9]) if pd.notna(row.iloc[9]) else 0.0,
                            'C': float(row.iloc[11]) if pd.notna(row.iloc[11]) else 0.0,
                            'D': float(row.iloc[13]) if pd.notna(row.iloc[13]) else 0.0
                        }
                        
                        kolom_mapping = {'A': 7, 'B': 9, 'C': 11, 'D': 13}
                        kunci_jawaban = None
                        for opsi, col_idx in kolom_mapping.items():
                            cell_obj = ws_rangkuman.cell(row=excel_row_num, column=col_idx)
                            cell_warna = str(cell_obj.fill.start_color.rgb).upper().strip()
                            if cell_obj.fill.fill_type is not None and cell_warna not in ['00000000', '000000', 'FFFFFFFF', 'NONE', 'INDEXED']:
                                kunci_jawaban = opsi
                                break
                        
                        if not kunci_jawaban:
                            kunci_jawaban = min(opsi_pct, key=lambda k: abs(opsi_pct[k] - mean_val))
                    except Exception:
                        continue

                    if max(opsi_pct.values()) == 0.0:
                        ws_rangkuman.cell(row=excel_row_num, column=15, value="Semua Distraktor tidak efektif")
                        continue

                    tidak_efektif, cukup_efektif, sangat_efektif, overpowered = [], [], [], []
                    for opsi, pct in opsi_pct.items():
                        if opsi == kunci_jawaban: continue
                        if pct > opsi_pct[kunci_jawaban]: overpowered.append(opsi)
                        elif mean_val >= 0.90 and pct > 0: cukup_efektif.append(opsi)
                        elif pct >= 0.10: sangat_efektif.append(opsi)
                        elif pct >= 0.05: cukup_efektif.append(opsi)
                        else: tidak_efektif.append(opsi)

                    kalimat_final = []
                    if overpowered: kalimat_final.append(f"Disatraktor {format_opsi_nama(overpowered)} sangat efektif bahkan cenderung dipilih dibanding kunci jawaban")
                    if sangat_efektif: kalimat_final.append(f"Distraktor {format_opsi_nama(sangat_efektif)} sangat efektif")
                    if cukup_efektif: kalimat_final.append(f"Distraktor {format_opsi_nama(cukup_efektif)} cukup efektif")
                    if tidak_efektif: kalimat_final.append(f"Distraktor {format_opsi_nama(tidak_efektif)} tidak efektif")

                    text_kesimpulan = "; ".join(kalimat_final) if overpowered else ", ".join(kalimat_final)
                    ws_rangkuman.cell(row=excel_row_num, column=15, value=text_kesimpulan)

            # --- PROSES 2: AUTOMATED GRAPHIC PLOTTER KHUSUS SHEET 3 (EVALUASI 1) ---
            with st.spinner("Sistem sedang menggambar Grafik Evaluasi 1 ke Sheet 3..."):
                df_clean = df_calc.dropna(subset=['No Item', 'Mean', 'Corrected Item-Total Correlation']).copy()
                df_clean = df_clean[df_clean['No Item'].str.contains('VAR', na=False)]
                
                # Strategi label ringkas (VAR00010 -> 10)
                def pangkas_label_item(teks_item):
                    ekstrak_angka = re.search(r'\d+', str(teks_item))
                    if ekstrak_angka:
                        return str(int(ekstrak_angka.group()))
                    return str(teks_item)
                
                df_clean['Label_Singkat'] = df_clean['No Item'].apply(pangkas_label_item)
                
                fig, ax = plt.subplots(figsize=(13, 10))
                
                # PENTING: Sumbu Horizontal = Mean, Sumbu Vertikal = CITC (Korelasi)
                # Sumbu X (Horizontal): Rentang 0.0 - 1.0 dengan interval 0.1
                # Sumbu Y (Vertikal): Rentang -0.5 - 1.0 dengan interval 0.1
                ax.set_xticks(np.arange(0.0, 1.1, 0.1))
                ax.set_yticks(np.arange(-0.5, 1.1, 0.1))
                
                # SYARAT EVALUASI 1: Balik arah horizontal dari KANAN KE KIRI (0 di kanan, 1 di kiri)
                ax.set_xlim(1.05, -0.05)
                ax.set_ylim(-0.55, 1.05)
                
                # Gambar Garis Bantu Batas Wilayah A, B, C, D, E, F secara visual
                # Batas Vertikal Utama (Horizontal X di Grafik)
                ax.axvline(x=0.3, color='black', linestyle=':', alpha=0.4)
                ax.axvline(x=0.7, color='black', linestyle=':', alpha=0.4)
                # Batas Horizontal Utama (Vertikal Y di Grafik)
                ax.axhline(y=0.3, color='black', linestyle=':', alpha=0.4)
                
                # Plot Titik Soal (Bintang Oranye)
                ax.scatter(df_clean['Mean'], df_clean['Corrected Item-Total Correlation'], 
                           color='#FF8C00', marker='*', s=160, edgecolor='black', linewidth=0.5, label='Butir Soal')
                
                # Labeling teks ringkas di atas titik
                for _, row in df_clean.iterrows():
                    ax.annotate(row['Label_Singkat'], 
                                (float(row['Mean']), float(row['Corrected Item-Total Correlation'])),
                                textcoords="offset points", 
                                xytext=(0, 6), 
                                ha='center', 
                                fontsize=8, 
                                fontweight='bold')
                
                # Teks Penempatan Wilayah Evaluasi 1 di Titik Tengah Masing-Masing Bidang
                # Wilayah A (0.3 - 0.7, 0.3 - 1.0)
                ax.text(0.5, 0.65, 'WILAYAH A', fontsize=12, fontweight='bold', color='darkgreen', ha='center', va='center', bbox=dict(facecolor='white', alpha=0.6, edgecolor='none'))
                # Wilayah B (0.0 - 0.3, 0.3 - 1.0)
                ax.text(0.15, 0.65, 'WILAYAH B', fontsize=12, fontweight='bold', color='navy', ha='center', va='center', bbox=dict(facecolor='white', alpha=0.6, edgecolor='none'))
                # Wilayah C (0.0 - 0.3, 0.0 - 0.3)
                ax.text(0.15, 0.15, 'WILAYAH C', fontsize=12, fontweight='bold', color='crimson', ha='center', va='center', bbox=dict(facecolor='white', alpha=0.6, edgecolor='none'))
                # Wilayah D (0.3 - 0.7, 0.0 - 0.3)
                ax.text(0.5, 0.15, 'WILAYAH D', fontsize=12, fontweight='bold', color='purple', ha='center', va='center', bbox=dict(facecolor='white', alpha=0.6, edgecolor='none'))
                # Wilayah E (0.7 - 1.0, 0.0 - 0.3)
                ax.text(0.85, 0.15, 'WILAYAH E', fontsize=12, fontweight='bold', color='darkorange', ha='center', va='center', bbox=dict(facecolor='white', alpha=0.6, edgecolor='none'))
                # Wilayah F (0.7 - 1.0, 0.3 - 1.0)
                ax.text(0.85, 0.65, 'WILAYAH F', fontsize=12, fontweight='bold', color='teal', ha='center', va='center', bbox=dict(facecolor='white', alpha=0.6, edgecolor='none'))
                
                # Dekorasi Grafik
                ax.set_title("PETA SEBARAN MATRIKS KUALITAS AITEM (EVALUASI 1 - REVERSE HORIZONTAL)", fontsize=13, fontweight='bold', pad=15)
                ax.set_xlabel("Tingkat Kesulitan (Mean) [Kanan ke Kiri]", fontsize=11, fontweight='bold', labelpad=10)
                ax.set_ylabel("Daya Beda (Corrected Item-Total Correlation)", fontsize=11, fontweight='bold', labelpad=10)
                ax.grid(True, linestyle=':', alpha=0.6)
                ax.legend(loc='lower left', fontsize=9)
                
                plt.tight_layout()
                
                # Simpan Grafik ke Objek Memori
                img_buf = io.BytesIO()
                plt.savefig(img_buf, format='png', dpi=180)
                img_buf.seek(0)
                plt.close(fig)
                
                # Masukkan ke Sheet 3
                if 'GRAFIK POSISI AITEM' in wb.sheetnames:
                    ws_sheet3 = wb['GRAFIK POSISI AITEM']
                    ws_sheet3._images.clear() 
                else:
                    ws_sheet3 = wb.create_sheet(title='GRAFIK POSISI AITEM')
                
                xl_img = OpenpyxlImage(img_buf)
                ws_sheet3.add_image(xl_img, 'B3')

            st.success("Sukses! Aturan Evaluasi 1 (Reverse Horizontal & Matriks Area A-F) Berhasil Diterapkan.")

            # 9. Export Gateway
            output = io.BytesIO()
            wb.save(output)
            processed_data = output.getvalue()

            st.download_button(
                label="📥 Download Excel Hasil Pemrosesan Final (Evaluasi 1)",
                data=processed_data,
                file_name="hasil_analisis_quiz_kls_A_EVALUASI_1.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error Deteksi Sistem: {e}. Pastikan format file Excel sesuai.")
